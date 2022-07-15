import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from twilio.rest import Client
from keys import *
from config import *


adv_urls = []


def send_sms(wiadomosc):
    client = Client(account_sid, auth_token)
    message = client.messages.create(
        body=wiadomosc,
        from_=twilio_number,
        to=my_phone_number
    )


def merge_rows(min_row, max_row):
    for j in range(min_row, max_row):
        # zakres
        sheet.merge_cells('B{0}:C{0}'.format(j))
        sheet.merge_cells('D{0}:H{0}'.format(j))
        sheet.merge_cells('I{0}:V{0}'.format(j))


def find_adv():
    for i in range(1, liczba_stron):
        url + str(i)
        page = requests.get(url)
        soup = BeautifulSoup(page.text, 'html.parser')
        ogloszenia = soup.find_all('a', href=True)
        for pages in ogloszenia:
            if pages['href'] not in adv_urls:
                if '/oferta/' in pages['href']:
                    adv_urls.append(pages['href'])
    adv_list = soup.find(
        "h1", {"data-testid": "results-heading"}).get_text(strip=True)
    print(adv_list)
    return adv_list


def main():
    global sheet
    adv_list = find_adv()
    print(str(len(adv_urls)))
    if str(len(adv_urls)) not in adv_list:
        adv_urls.clear()
        find_adv()

    print(str(len(adv_urls)))

    print('Znaleziono:', len(adv_urls), 'ogłoszeń')

    try:
        wb = load_workbook(filename)
        sheet = wb.active
    except BaseException:
        wb = Workbook()
        sheet = wb.active
        merge_rows(min_row, max_row)

        sheet['BX1000'].value = 3
        sheet['A1'].value = "Lp."
        sheet['B1'].value = "ID"
        sheet['D1'].value = "Tytuł"
        sheet['I1'].value = "Link"

    max = sheet['BX1000'].value
    x = 1

    print("Sprawdzanie ogłoszeń...")
    check_version = "ttid" or "1.9 ttid" or "1.9ttid"
    did_combi = "kombi" or "combi" or "cabriolet" or "kabriolet" or "cabrio" or "Kabriolet"

    count_new = 0
    for elements in adv_urls:

        url = elements
        page = requests.get(url)
        soup = BeautifulSoup(page.text, 'html.parser')

        try:
            version = soup.find(
                text="Wersja").next_element.next_element.get_text(
                strip=True).lower()
        except BaseException:
            version = ' '

        try:
            KM = soup.find(
                text="Moc").next_element.next_element.get_text().strip()
        except BaseException:

            KM = ' '

        try:
            desc = soup.find(
                text="Opis").next_element.next_element.get_text().strip()
        except BaseException:
            desc = ' '

        try:
            title = soup.find(class_="tags").get_text().strip()
        except BaseException:
            title = ' '

        if (KM == "180 KM" or KM == "210 KM") and (
                check_version in version or check_version in desc or check_version in title):

            try:
                ID = soup.find(id="ad_id").get_text().strip()
            except BaseException:
                pass
            try:
                body_type = soup.find(
                    text="Typ nadwozia").next_element.next_element.get_text(
                    strip=True).lower()

            except BaseException:
                body_type = " "

            if did_combi not in desc or did_combi not in title or did_combi not in body_type or did_combi not in version:
                for i in range(2, max):
                    if sheet.cell(
                            row=i, column=2).value is None and sheet.cell(
                            row=i, column=1).value is None:
                        if i == max - 1:
                            merge_rows(i, i + 1)
                        sheet.cell(row=i, column=9).value = elements  # link
                        sheet.cell(row=i, column=1).value = i - 1  # Lp
                        sheet.cell(row=i, column=2).value = ID  # ID
                        sheet.cell(
                            row=i, column=4).value = "------------"  # title
                        max = sheet['BX1000'].value = max + 1
                        count_new = count_new + 1
                        send_sms(title_msg + elements)
                        break

                    elif sheet.cell(row=i, column=2).value == ID:
                        break
                x = x + 1
            else:
                pass

    print("Znaleziono {} wyników spełniających warunki.".format(x - 1))

    if count_new == 0:
        print("Niestety nie znaleziono nowych ogłoszeń")
    else:
        print("Znaleziono {0} nowy/ch wyników".format(count_new))

    print("Zapisywanie działań...")
    wb.save(filename)
    print('Zapisano plik Excel!')


main()
